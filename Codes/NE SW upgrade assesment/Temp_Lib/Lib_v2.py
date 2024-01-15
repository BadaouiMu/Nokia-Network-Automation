import pandas as pd
import camelot
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import paramiko
from socket import error as socket_error
import warnings
from datetime import datetime

# Get the current date and time



def test_ssh(host, username, password):
    warnings.filterwarnings("ignore")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(host, username=username, password=password, timeout=10)
        return True
    except paramiko.AuthenticationException:
        return False
    except paramiko.SSHException as ssh_ex:
        return False
    except socket_error:
        return False
    finally:
        ssh.close()
    return False

def test_ssh_jump(host, username, password, jumphost, jumpuser, jumppass):
    warnings.filterwarnings("ignore")
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(jumphost, username=jumpuser, password=jumppass, timeout=5)
        transport = ssh.get_transport()
        dest_addr = (host, 22)
        local_addr = ('', 0)
        channel = transport.open_channel("direct-tcpip", dest_addr, local_addr)
        target_ssh = paramiko.SSHClient()
        target_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target_ssh.connect(host, username=username, password=password, sock=channel, timeout=5)
        return True
    except paramiko.AuthenticationException:
        return False
    except paramiko.SSHException as ssh_ex:
        return False
    except socket_error:
        return False
    finally:
        ssh.close()
    return False

def read_text(path):
    warnings.filterwarnings("ignore")
    list=[]

    with open(path) as device_file:
        
        for line in device_file:
            if line and line != "" and "," in line:
                list1=[]
                ipaddr, username, password = line.strip().split(",")
                list1.append(ipaddr)
                list1.append(username)
                list1.append(password)
                list.append(list1)

    return list

def update_csv_file(ip_address,system, version, list1, list2):
    warnings.filterwarnings("ignore")
    try:
        df = pd.read_csv("network_inventory.csv")
    except FileNotFoundError:
        df = pd.DataFrame(columns=["IP Address","System Name", "Current Version"])

    mask = df["IP Address"] == ip_address
    if mask.any():
        df = df[~mask]
        df.reset_index(drop=True, inplace=True)
    
    version_name = version[0][0] if len(version) >= 1 and len(version[0]) == 1 else "not found"
    system_name = system[0][0] if len(system) == 1 and len(system[0]) == 1 else "not found"

    row_dict = {"IP Address": ip_address,"System Name":system_name, "Current Version":version_name}
    for column, value1, value2 in list1:
        column1= "Card " + column + " Name"
        row_dict[column1]=value1
        column2= "Card " + column + " Part N"
        value2 = value2[:10]
        row_dict[column2] = value2
    for column, value1, value2 in list2:
        column1= "Card " + column + " Name"
        row_dict[column1]=value1
        column2= "Card " + column + " Part N"
        value2 = value2[:10]
        row_dict[column2] = value2
    df = df.append(row_dict, ignore_index=True)
    
    cols_to_sort = sorted([col for col in df.columns if col not in ["IP Address", "System Name", "Current Version"]])
    sorted_cols = ["IP Address", "System Name", "Current Version"] + cols_to_sort
    df = df[sorted_cols]

    df.to_csv("network_inventory.csv", index=False)
    

def extract_data_from_pdf( pdf_file_path, chapitre):
    warnings.filterwarnings("ignore")
    try:
        df = pd.read_csv(f"{chapitre}.csv")
    except FileNotFoundError:
        columns_to_extract = ["Nokia Part #", "Description", "sat-type", "CLI String (MDA)", "CLI String (Card)"]

        pdf_reader = PyPDF2.PdfReader(open(pdf_file_path, 'rb'))  
        page_numbers = []
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]  
            if chapitre in page.extract_text():
                page_numbers.append(page_num + 1)

        tables_to_concatenate = []
        for page_num in page_numbers:
            tables = camelot.read_pdf(pdf_file_path, pages=str(page_num))
            for table in tables:
                if any("Nokia Part #" in col for col in table.df.loc[0].tolist()):
                    extracted_table=table.df
                    extracted_table.loc[0] = extracted_table.loc[0].apply(lambda x: x.replace('\n', '') if isinstance(x, str) else x)
                    extracted_table.columns = table.df.iloc[0]
                    extracted_table = extracted_table.iloc[1:]
                    extracted_table.drop([col for col in extracted_table.columns.tolist() if col not in columns_to_extract], axis=1, inplace=True)
                    tables_to_concatenate.append(extracted_table)

        big_table = pd.concat(tables_to_concatenate)
        big_table.to_csv(f"{chapitre}.csv", index=False)


def compare_csv_file(chapitre):  
    warnings.filterwarnings("ignore")
    df1 = pd.read_csv('network_inventory.csv')
    df2 = pd.read_csv(f"{chapitre}.csv")
    
    wb = Workbook()
    ws = wb.active
    
    headers = list(df1.columns)
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])
        
    for col in range(1, len(df1.columns)+1):
        ws.column_dimensions[chr(64+col)].width = 20
        
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')

    for row in range(2, len(df1)+2):
        for col in range(1, len(headers)+1):
            if col <= 3:  
                value = df1.iloc[row-2, col-1]
                cell = ws.cell(row=row, column=col, value=value)
            elif col % 2 == 1 :
                value = df1.iloc[row-2, col-1]
                t=0 
                for nokia_part in df2['Nokia Part #'].values:
                    if len(nokia_part) == 10 and value == nokia_part:
                        #print('here')
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.fill = green_fill
                        t=1
                        break
                    elif type(value)==float:
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.fill = blue_fill
                        t=2
                    elif len(nokia_part) != 10:
                        part = nokia_part.split("\n")
                        first_half = part[0]
                        second_half = part[1]

                        if value == first_half or value == second_half:
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.fill = green_fill
                            t=1
                            break
                        else:
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.fill = red_fill
                prev_col = col - 1
                prev_value = df1.iloc[row-2, prev_col-1]
                prev_cell = ws.cell(row=row, column=prev_col, value=prev_value)
                if t==0:
                    prev_cell.fill= red_fill
                elif t==1:
                    prev_cell.fill= green_fill
                else:
                    prev_cell.fill= blue_fill
            #elif col % 2 == 0:
                #value = df1.iloc[row-2, col-1]
                #for name in df2[["sat-type", "CLI String (MDA)", "CLI String (Card)"]].values:
                    #if value in name:
                        ##print('here')
                        #cell = ws.cell(row=row, column=col, value=value)
                        #cell.fill = green_fill
                        #t=1
                        #break
                    #elif type(value)==float:
                        #cell = ws.cell(row=row, column=col, value=value)
                        #cell.fill = blue_fill
                        #t=2
                    #else:
                            #cell = ws.cell(row=row, column=col, value=value)
                            #cell.fill = red_fill                

                
    wb.save('output_from_routers.xlsx')

def compare_excel_file(path, chapitre):
    warnings.filterwarnings("ignore")

    #df1 = pd.read_csv("/home/sats/workspace/user_robots/Mus/network_inventory/tmp/excell.csv")
    df1 = pd.read_excel(f"{path}")
    df2 = pd.read_csv(f"{chapitre}.csv")



    # create a new Excel workbook
    workbook = Workbook()

    # create a worksheet in the workbook and name it
    worksheet = workbook.active
    worksheet.title = 'Comparison'


    headers = list(df1.columns)
    for col in range(len(headers)):
        worksheet.cell(row=1, column=col+1, value=headers[col])

    # define the fill colors for the rows
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')


    # loop through each row in the first CSV file
    for index, row in df1.iterrows():
        # get the value in the "Part Number" column
        part_number = row['Part Number']
        if type(part_number) != float and len(part_number) > 10:
            part_number= part_number[:10]

        # check if the part number exists in the second CSV file
        if type(part_number)==float or len(part_number) < 9:
            row_color = blue_fill
        elif part_number in df2['Nokia Part #'].values:
            # set the row color to green
            row_color = green_fill

        else:
            # set the row color to red
            row_color = red_fill

        # write the row to the Excel worksheet with the appropriate color
        # write the row to the Excel worksheet with the appropriate color
        row_num = index + 2
        for col_num, value in enumerate(row.to_list(), 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.fill = row_color

    # save the Excel workbook
    now = datetime.now()
    date_time = now.strftime("%d_%m_%Y_%H_%M_%S")
    workbook.save(f'output_from_excel_{date_time}.xlsx')